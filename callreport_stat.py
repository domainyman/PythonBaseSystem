# -*- coding: utf-8 -*-

import decimal
import sys

from PyQt5.QtCore import QDate
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QMessageBox
from openpyxl import load_workbook
from connectdb import database
from statement_reportui import Ui_reportui


class Callreport(QMainWindow, Ui_reportui):
    def __init__(self, parent=None):
        super(Callreport, self).__init__(parent)
        self.setupUi(self)
        decimal.getcontext().rounding = "ROUND_HALF_UP"
        self.condb = database()
        self.end_ed.setDate(QDate.currentDate())
        self.start_ed.setDate(QDate.currentDate())
        self.report_bu()

    def search_dl(self):
        wb = load_workbook('./tmp/Report.xlsx')
        ws = wb['Report']
        ws['A1'] = "Monthly Financial Report" + self.start_ed.text() + "-" + self.end_ed.text()
        da_st = self.start_ed.text()
        da_du = self.end_ed.text()

        # Item_inv
        self.re_da, self.re_iv, self.re_ci, self.re_am, self.re_du, self.re_pay = self.condb.report_inv_dl(da_st, da_du)
        print(self.re_da)
        a = 6
        num_inv = 1
        for val_inv_nu in range(int(len(self.re_iv))):
            ws.cell(a, 1).value = num_inv
            a = a + 1
            num_inv = num_inv + 1
            print(num_inv)
        b = 6
        for val_da in range(int(len(self.re_da))):
            ws.cell(b, 2).value = self.re_da[val_da]
            b = b + 1
            print(b)
        c = 6
        for val_iv in range(int(len(self.re_iv))):
            ws.cell(c, 3).value = self.re_iv[val_iv]
            c = c + 1
            print(c)
        d = 6
        for val_ci in range(int(len(self.re_ci))):
            ws.cell(d, 4).value = self.re_ci[val_ci]
            d = d + 1
            print(d)
        e = 6
        for val_am in range(int(len(self.re_am))):
            ws.cell(e, 5).value = decimal.Decimal(str((self.re_am[val_am])))
            e = e + 1
            print(e)
        f = 6
        for val_du in range(int(len(self.re_du))):
            ws.cell(f, 6).value = self.re_du[val_du]
            f = f + 1
            print(f)
        g = 6
        for val_pay in range(int(len(self.re_pay))):
            ws.cell(g, 7).value = self.re_pay[val_pay]
            g = g + 1
            print(g)

        # Item_exp
        self.reex_da, self.reex_iv, self.reex_sh, self.reex_am = self.condb.report_exp_dl(da_st, da_du)
        print(self.reex_da)
        h = 6
        numex_inv = 1
        for valex_inv_nu in range(int(len(self.reex_iv))):
            ws.cell(h, 8).value = numex_inv
            h = h + 1
            numex_inv = numex_inv + 1
            print(numex_inv)
        i = 6
        for valex_da in range(int(len(self.reex_da))):
            ws.cell(i, 9).value = self.reex_da[valex_da]
            i = i + 1
            print(i)
        j = 6
        for valex_iv in range(int(len(self.reex_iv))):
            ws.cell(j, 10).value = self.reex_iv[valex_iv]
            j = j + 1
            print(j)
        k = 6
        for valex_sh in range(int(len(self.reex_sh))):
            ws.cell(k, 11).value = self.reex_sh[valex_sh]
            k = k + 1
            print(k)
        la = 6
        for valex_am in range(int(len(self.reex_am))):
            ws.cell(la, 12).value = decimal.Decimal(str((self.reex_am[valex_am])))
            la = la + 1
            print(la)

        # Item_pur
        self.repu_da, self.repu_iv, self.repu_su, self.repu_am = self.condb.report_pur_dl(da_st, da_du)
        print(self.repu_da)
        m = 6
        numpu_inv = 1
        for valpu_inv_nu in range(int(len(self.repu_iv))):
            ws.cell(m, 13).value = numpu_inv
            m = m + 1
            numpu_inv = numpu_inv + 1
            print(numpu_inv)
        n = 6
        for valpu_da in range(int(len(self.repu_da))):
            ws.cell(n, 14).value = self.repu_da[valpu_da]
            n = n + 1
            print(n)
        o = 6
        for valpu_iv in range(int(len(self.repu_iv))):
            ws.cell(o, 15).value = self.repu_iv[valpu_iv]
            o = o + 1
            print(o)
        p = 6
        for valpu_sh in range(int(len(self.repu_su))):
            ws.cell(p, 16).value = self.repu_su[valpu_sh]
            p = p + 1
            print(p)
        q = 6
        for valpu_am in range(int(len(self.repu_am))):
            ws.cell(q, 17).value = decimal.Decimal(str((self.repu_am[valpu_am])))
            q = q + 1
            print(q)

        # Item_cn
        self.recn_da, self.recn_iv, self.recn_sh, self.recn_am, self.recn_clear = self.condb.report_cn_dl(da_st, da_du)
        print(self.repu_da)
        r = 6
        numcn_inv = 1
        for valcn_inv_nu in range(int(len(self.recn_iv))):
            ws.cell(r, 18).value = numcn_inv
            r = r + 1
            numcn_inv = numcn_inv + 1
            print(numcn_inv)
        s = 6
        for valcn_da in range(int(len(self.recn_da))):
            ws.cell(s, 19).value = self.recn_da[valcn_da]
            s = s + 1
            print(s)
        t = 6
        for valcn_iv in range(int(len(self.recn_iv))):
            ws.cell(t, 20).value = self.recn_iv[valcn_iv]
            t = t + 1
            print(t)
        u = 6
        for valcn_sh in range(int(len(self.recn_sh))):
            ws.cell(u, 21).value = self.recn_sh[valcn_sh]
            u = u + 1
            print(u)
        v = 6
        for valcn_am in range(int(len(self.recn_am))):
            ws.cell(v, 22).value = decimal.Decimal(str(self.recn_am[valcn_am]))
            v = v + 1
            print(v)
        w = 6
        for valcn_clear in range(int(len(self.recn_clear))):
            ws.cell(w, 23).value = self.recn_clear[valcn_clear]
            w = w + 1
            print(w)

        # Item_return
        self.reret_da, self.reret_iv, self.reret_sh, self.reret_am = self.condb.report_ret_dl(da_st, da_du)
        print(self.reret_da)
        x = 6
        numret_inv = 1
        for valret_inv_nu in range(int(len(self.reret_iv))):
            ws.cell(x, 24).value = numret_inv
            x = x + 1
            numret_inv = numret_inv + 1
            print(numret_inv)
        y = 6
        for valret_da in range(int(len(self.reret_da))):
            ws.cell(y, 25).value = self.reret_da[valret_da]
            y = y + 1
            print(y)
        z = 6
        for valret_iv in range(int(len(self.reret_iv))):
            ws.cell(z, 26).value = self.reret_iv[valret_iv]
            z = z + 1
            print(z)
        aa = 6
        for valret_sh in range(int(len(self.reret_sh))):
            ws.cell(aa, 27).value = self.reret_sh[valret_sh]
            aa = aa + 1
            print(aa)
        ab = 6
        for valret_am in range(int(len(self.reret_am))):
            ws.cell(ab, 28).value = decimal.Decimal(str(self.reret_am[valret_am]))
            ab = ab + 1
            print(ab)

        self.dirpath, _ = QFileDialog.getSaveFileName(self, "Save file",
                                                      "./" + "Monthly Financial Report" + self.start_ed.text() + "-" + self.end_ed.text(),
                                                      'xlsx(*.xlsx)')
        if self.dirpath != "":
            try:
                with open(self.dirpath, 'w'):
                    print(self.dirpath)
                    wb.save(self.dirpath)
                    QMessageBox.about(self, "About", "Done. !", )
            except IOError:
                QMessageBox.about(self, "About", "Pleare Close Document or Change Document Name. !", )
            else:
                pass
        else:
            QMessageBox.about(self, "About", "EXIT. !", )

    def report_bu(self):
        self.Search_bu.clicked.connect(self.search_dl)

    def closeEvent(self, event):
        self.condb.closedb()
        print('window close')


if __name__ == "__main__":
    app = QApplication(sys.argv)
    win = Callreport()
    win.show()
    sys.exit(app.exec_())
