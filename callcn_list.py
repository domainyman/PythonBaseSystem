# -*- coding: utf-8 -*-

import sys

from PyQt5 import QtCore
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QMessageBox
from cn_listui import Ui_cn_list
from connectdb import database
from PyQt5.QtCore import Qt
from openpyxl import load_workbook
import decimal


class callinv_list(QMainWindow, Ui_cn_list):
    my_Signal = QtCore.pyqtSignal()

    def __init__(self, list_info, parent=None):
        super(callinv_list, self).__init__(parent)
        self.data_info = list_info
        self.data_inv = self.data_info[1]
        decimal.getcontext().rounding = "ROUND_HALF_UP"
        self.setupUi(self)
        self.clea_box.addItems(["No", "Yes"])
        self.condb = database()
        self.info = self.condb.display_cominfo()
        self.setWindowModality(Qt.ApplicationModal)
        self.tolist()
        self.bu_click()

    def tolist(self):
        self.cn_li.setText(self.data_info[1])
        self.da_li.setText(self.data_info[2])
        self.billt_li.setText(self.data_info[3])
        self.addr_li.setText(self.data_info[4])
        self.attn_li.setText(self.data_info[5])
        self.tel_lu.setText(self.data_info[6])
        self.cusid_li.setText(self.data_info[7])
        self.dued_li.setText(self.data_info[8])
        self.sales_li.setText(self.data_info[9])
        self.saleste_li.setText(self.data_info[10])
        self.p1.setText(self.data_info[11])
        self.p1m.setText(self.data_info[12])
        self.p1c.setText(self.data_info[13])
        self.p1p.setText(self.data_info[14])
        self.p2.setText(self.data_info[15])
        self.p2m.setText(self.data_info[16])
        self.p2c.setText(self.data_info[17])
        self.p2p.setText(self.data_info[18])
        self.p3.setText(self.data_info[19])
        self.p3m.setText(self.data_info[20])
        self.p3c.setText(self.data_info[21])
        self.p3p.setText(self.data_info[22])
        self.p4.setText(self.data_info[23])
        self.p4m.setText(self.data_info[24])
        self.p4c.setText(self.data_info[25])
        self.p4p.setText(self.data_info[26])
        self.p5.setText(self.data_info[27])
        self.p5m.setText(self.data_info[28])
        self.p5c.setText(self.data_info[29])
        self.p5p.setText(self.data_info[30])
        self.p6.setText(self.data_info[31])
        self.p6m.setText(self.data_info[32])
        self.p6c.setText(self.data_info[33])
        self.p6p.setText(self.data_info[34])
        self.p7.setText(self.data_info[35])
        self.p7m.setText(self.data_info[36])
        self.p7c.setText(self.data_info[37])
        self.p7p.setText(self.data_info[38])
        self.p8.setText(self.data_info[39])
        self.p8m.setText(self.data_info[40])
        self.p8c.setText(self.data_info[41])
        self.p8p.setText(self.data_info[42])
        self.p9.setText(self.data_info[43])
        self.p9m.setText(self.data_info[44])
        self.p9c.setText(self.data_info[45])
        self.p9p.setText(self.data_info[46])
        self.p10.setText(self.data_info[47])
        self.p10m.setText(self.data_info[48])
        self.p10c.setText(self.data_info[49])
        self.p10p.setText(self.data_info[50])
        self.c1.setText(self.data_info[51])
        self.c2.setText(self.data_info[52])
        self.c3.setText(self.data_info[53])
        self.am_ed.setText(self.data_info[54])
        self.clear_ed.setText(self.data_info[55])

    def upload_clear(self):
        reply = QMessageBox.information(self, "Upload CN Note", "Once Upload , Invoice cannot be recovered!",
                                        QMessageBox.Yes | QMessageBox.Close, QMessageBox.Close)
        if reply == QMessageBox.Yes:
            iv = self.cn_li.text()
            pay_st = self.clea_box.currentText()
            self.condb.cnuplpaym_db(iv, pay_st)
            self.close()

    def del_inv_bu(self):
        reply = QMessageBox.information(self, "Delete CN Note", "Once deleted , Invoice cannot be recovered!",
                                        QMessageBox.Yes | QMessageBox.Close, QMessageBox.Close)
        if reply == QMessageBox.Yes:
            self.iv = self.data_inv
            self.condb.del_cn_db(self.iv)
            self.close()


    def xls(self):
        wb = load_workbook('./tmp/Sample.xlsx')
        ws = wb['Invoice']
        ws['F1'] = "Credit Note"
        ws['A1'] = self.info[1]
        ws['B2'] = self.info[2]
        ws['B3'] = self.sales_li.text()
        ws['B4'] = self.saleste_li.text()
        ws['B8'] = self.billt_li.text()
        ws['B9'] = self.addr_li.text()
        ws['B10'] = self.attn_li.text()
        ws['B11'] = self.tel_lu.text()
        ws['G3'] = self.da_li.text()
        ws['G4'] = self.cn_li.text()
        ws['G5'] = self.cusid_li.text()
        ws['G6'] = self.dued_li.text()
        ws['A16'] = self.p1m.text()
        ws['A17'] = self.p2m.text()
        ws['A18'] = self.p3m.text()
        ws['A19'] = self.p4m.text()
        ws['A20'] = self.p5m.text()
        ws['A21'] = self.p6m.text()
        ws['A22'] = self.p7m.text()
        ws['A23'] = self.p8m.text()
        ws['A24'] = self.p9m.text()
        ws['A25'] = self.p10m.text()
        ws['B16'] = self.p1.text()
        ws['B17'] = self.p2.text()
        ws['B18'] = self.p3.text()
        ws['B19'] = self.p4.text()
        ws['B20'] = self.p5.text()
        ws['B21'] = self.p6.text()
        ws['B22'] = self.p7.text()
        ws['B23'] = self.p8.text()
        ws['B24'] = self.p9.text()
        ws['B25'] = self.p10.text()
        ws['A41'] = self.c1.text()
        ws['A42'] = self.c2.text()
        ws['A43'] = self.c3.text()

        ###Product
        self.toam = []
        self.toqty = []
        if self.p1c.text() != "":
            ws['E16'] = decimal.Decimal(self.p1c.text())
            ws['F16'] = decimal.Decimal(self.p1p.text())
            ws['G16'] = decimal.Decimal(self.p1p.text()) * decimal.Decimal(self.p1c.text())
            self.toam.append(ws['G16'].value)
            self.toqty.append(ws['E16'].value)
        if self.p2c.text() != "":
            ws['E17'] = decimal.Decimal(self.p2c.text())
            ws['F17'] = decimal.Decimal(self.p2p.text())
            ws['G17'] = decimal.Decimal(self.p2p.text()) * decimal.Decimal(self.p2c.text())
            self.toam.append(ws['G17'].value)
            self.toqty.append(ws['E17'].value)
        if self.p3c.text() != "":
            ws['E18'] = decimal.Decimal(self.p3c.text())
            ws['F18'] = decimal.Decimal(self.p3p.text())
            ws['G18'] = decimal.Decimal(self.p3p.text()) * decimal.Decimal(self.p3c.text())
            self.toam.append(ws['G18'].value)
            self.toqty.append(ws['E18'].value)
        if self.p4c.text() != "":
            ws['E19'] = decimal.Decimal(self.p4c.text())
            ws['F19'] = decimal.Decimal(self.p4p.text())
            ws['G19'] = decimal.Decimal(self.p4p.text()) * decimal.Decimal(self.p4c.text())
            self.toam.append(ws['G19'].value)
            self.toqty.append(ws['E19'].value)
        if self.p5c.text() != "":
            ws['E20'] = decimal.Decimal(self.p5c.text())
            ws['F20'] = decimal.Decimal(self.p5p.text())
            ws['G20'] = decimal.Decimal(self.p5p.text()) * decimal.Decimal(self.p5c.text())
            self.toam.append(ws['G20'].value)
            self.toqty.append(ws['E20'].value)
        if self.p6c.text() != "":
            ws['E21'] = decimal.Decimal(self.p6c.text())
            ws['F21'] = decimal.Decimal(self.p6p.text())
            ws['G21'] = decimal.Decimal(self.p6p.text()) * decimal.Decimal(self.p6c.text())
            self.toam.append(ws['G21'].value)
            self.toqty.append(ws['E21'].value)
        if self.p7c.text() != "":
            ws['E22'] = decimal.Decimal(self.p7c.text())
            ws['F22'] = decimal.Decimal(self.p7p.text())
            ws['G22'] = decimal.Decimal(self.p7p.text()) * decimal.Decimal(self.p7c.text())
            self.toam.append(ws['G22'].value)
            self.toqty.append(ws['E22'].value)
        if self.p8c.text() != "":
            ws['E23'] = decimal.Decimal(self.p8c.text())
            ws['F23'] = decimal.Decimal(self.p8p.text())
            ws['G23'] = decimal.Decimal(self.p8p.text()) * decimal.Decimal(self.p8c.text())
            self.toam.append(ws['G23'].value)
            self.toqty.append(ws['E23'].value)
        if self.p9c.text() != "":
            ws['E24'] = decimal.Decimal(self.p9c.text())
            ws['F24'] = decimal.Decimal(self.p9p.text())
            ws['G24'] = decimal.Decimal(self.p9p.text()) * decimal.Decimal(self.p9c.text())
            self.toam.append(ws['G24'].value)
            self.toqty.append(ws['E24'].value)
        if self.p10c.text() != "":
            ws['E25'] = decimal.Decimal(self.p10c.text())
            ws['F25'] = decimal.Decimal(self.p10p.text())
            ws['G25'] = decimal.Decimal(self.p10p.text()) * decimal.Decimal(self.p10c.text())
            self.toam.append(ws['G25'].value)
            self.toqty.append(ws['E25'].value)
        print(self.toam)
        ws['G38'] = sum(self.toam)
        print(self.toqty)
        ws['E38'] = sum(self.toqty)
        self.dirpath, _ = QFileDialog.getSaveFileName(self, "Save file", "./" + self.cn_li.text(), 'xlsx(*.xlsx)')
        if self.dirpath != "":
            try:
                with open(self.dirpath, 'w'):
                    print(self.dirpath)
                    wb.save(self.dirpath)
                    QMessageBox.about(self, "About", "Done. !", )
            except IOError:
                QMessageBox.about(self, "About", "Pleare Close Document or Change Document Name. !", )
            else:
                pass
        else:
            QMessageBox.about(self, "About", "EXIT. !", )


    def checkBoxChangedAction(self, state):
        if state == Qt.Checked:
            self.da_li.setEnabled(True)
            self.dued_li.setEnabled(True)
        else:
            self.da_li.setEnabled(False)
            self.dued_li.setEnabled(False)

    def da_up_bu(self):
        reply = QMessageBox.information(self, "UPLOAD Upload", "Once UPLOAD , Upload cannot be recovered!",
                                        QMessageBox.Yes | QMessageBox.Close, QMessageBox.Close)
        if reply == QMessageBox.Yes:
            da = self.da_li.text()
            dua = self.dued_li.text()
            inv = self.cn_li.text()
            self.condb.up_cn_data(da,dua,inv)
            self.close()

    def bu_click(self):
        self.del_bu.clicked.connect(self.del_inv_bu)
        self.print_bu.clicked.connect(self.xls)
        self.upload_bu.clicked.connect(self.upload_clear)
        self.data_checkBox.stateChanged.connect(self.checkBoxChangedAction)
        self.data_upload.clicked.connect(self.da_up_bu)

    def sendEditContent(self):
        self.my_Signal.emit()

    def closeEvent(self, event):
        self.sendEditContent()
        self.condb.closedb()
        print('window close')


if __name__ == "__main__":
    app = QApplication(sys.argv)
    win = callinv_list()
    win.show()
    sys.exit(app.exec_())
