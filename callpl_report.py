# -*- coding: utf-8 -*-

import decimal
import sys
from PyQt5.QtCore import QDate
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QMessageBox
from openpyxl import load_workbook
import calendar
import datetime
from datetime import timedelta
from connectdb import database
from pl_reportui import Ui_plreportui


class Callplreport(QMainWindow, Ui_plreportui):
    def __init__(self, parent=None):
        super(Callplreport, self).__init__(parent)
        self.setupUi(self)
        self.timeset()
        decimal.getcontext().rounding = "ROUND_HALF_UP"
        self.condb = database()
        self.reportpl_bu()

    def timeset(self):
        self.start_ed_list = [self.start_ed, self.start_ed_2, self.start_ed_3, self.start_ed_4, self.start_ed_5,
                              self.start_ed_6,
                              self.start_ed_7, self.start_ed_8, self.start_ed_9, self.start_ed_10, self.start_ed_11,
                              self.start_ed_12]
        self.end_ed_list = [self.end_ed, self.end_ed_2, self.end_ed_3, self.end_ed_4, self.end_ed_5, self.end_ed_6,
                            self.end_ed_7, self.end_ed_8, self.end_ed_9, self.end_ed_10, self.end_ed_11, self.end_ed_12]
        self.start = []
        self.end = []
        now = datetime.date.today()
        this_month_start = datetime.date(now.year, now.month, 1)
        this_month_start_list = this_month_start.strftime('%Y-%m-%d')
        this_month_end = datetime.date(now.year, now.month, calendar.monthrange(now.year, now.month)[1])
        this_month_end_list = this_month_end.strftime('%Y-%m-%d')
        self.start.append(this_month_start_list)
        self.end.append(this_month_end_list)
        for x in range(11):
            last_month_1_end = this_month_start - timedelta(days=1)
            last_month_1_start = datetime.date(last_month_1_end.year, last_month_1_end.month, 1)
            # print(last_month_1_end)
            # print(last_month_1_start)
            this_month_start = last_month_1_start
            this_month_end = last_month_1_end
            # print(this_month_start)
            # print(this_month_end)
            this_month_start_list = this_month_start.strftime('%Y-%m-%d')
            this_month_end_list = this_month_end.strftime('%Y-%m-%d')
            self.start.append(this_month_start_list)
            self.end.append(this_month_end_list)
        self.start.reverse()
        self.end.reverse()
        print(self.start)
        print(self.end)
        for x in range(12):
            self.start_ed_list[x].setDate(QDate.fromString(self.start[x], 'yyyy-MM-dd'))
        for x in range(12):
            self.end_ed_list[x].setDate(QDate.fromString(self.end[x], 'yyyy-MM-dd'))

    def search_dl(self):
        wb = load_workbook('./tmp/Pl.xlsx')
        ws = wb['Statement']
        ws['A1'] = "Profìt and Loss Statement" + self.start_ed.text() + "-" + self.end_ed.text()
        self.item = ["Shipping Fee", "Exchange", "Bank Charge", "Stationery Supplies", "Insurance",
                     "Other Expenses", "Salary and Wages", "MPF", "Rent"]
        self.item_bankint = "Bank interest"
        # Item_1
        b = 5
        da_st = self.start_ed.text()
        da_du = self.end_ed.text()
        ws.cell(2, 2).value = da_st + "-" + da_du
        self.inv_n1 = self.condb.pl_inv_rep(da_st, da_du)
        ws.cell(3, 2).value = self.inv_n1
        self.pur_n1 = self.condb.pl_pur_rep(da_st, da_du)
        ws.cell(4, 2).value = self.pur_n1
        for n1 in range(len(self.item)):
            self.info = self.condb.pl_exp_rep(da_st, da_du, self.item[n1])
            ws.cell(b, 2).value = self.info
            b = b + 1
        self.cn_n1 = self.condb.pl_cn_rep(da_st, da_du)
        ws.cell(14, 2).value = self.cn_n1
        self.bkin = self.condb.pl_exp_rep(da_st, da_du, self.item_bankint)
        ws.cell(16, 2).value = self.bkin
        # Item_2
        c = 5
        da_st_2 = self.start_ed_2.text()
        da_du_2 = self.end_ed_2.text()
        ws.cell(2, 3).value = da_st_2 + "-" + da_du_2
        self.inv_n2 = self.condb.pl_inv_rep(da_st_2, da_du_2)
        ws.cell(3, 3).value = self.inv_n2
        self.pur_n2 = self.condb.pl_pur_rep(da_st_2, da_du_2)
        ws.cell(4, 3).value = self.pur_n2
        for n2 in range(len(self.item)):
            self.info2 = self.condb.pl_exp_rep(da_st_2, da_du_2, self.item[n2])
            ws.cell(c, 3).value = self.info2
            c = c + 1
        self.cn_n2 = self.condb.pl_cn_rep(da_st_2, da_du_2)
        ws.cell(14, 3).value = self.cn_n2
        self.bkin_n2 = self.condb.pl_exp_rep(da_st_2, da_du_2, self.item_bankint)
        ws.cell(16, 3).value = self.bkin_n2
        # Item_3
        d = 5
        da_st_3 = self.start_ed_3.text()
        da_du_3 = self.end_ed_3.text()
        ws.cell(2, 4).value = da_st_3 + "-" + da_du_3
        self.inv_n3 = self.condb.pl_inv_rep(da_st_3, da_du_3)
        ws.cell(3, 4).value = self.inv_n3
        self.pur_n3 = self.condb.pl_pur_rep(da_st_3, da_du_3)
        ws.cell(4, 4).value = self.pur_n3
        for n3 in range(len(self.item)):
            self.info3 = self.condb.pl_exp_rep(da_st_3, da_du_3, self.item[n3])
            ws.cell(d, 4).value = self.info3
            d = d + 1
        self.cn_n3 = self.condb.pl_cn_rep(da_st_3, da_du_3)
        ws.cell(14, 4).value = self.cn_n3
        self.bkin_n3 = self.condb.pl_exp_rep(da_st_3, da_du_3, self.item_bankint)
        ws.cell(16, 4).value = self.bkin_n3
        # Item_4
        e = 5
        da_st_4 = self.start_ed_4.text()
        da_du_4 = self.end_ed_4.text()
        ws.cell(2, 5).value = da_st_4 + "-" + da_du_4
        self.inv_n4 = self.condb.pl_inv_rep(da_st_4, da_du_4)
        ws.cell(3, 5).value = self.inv_n4
        self.pur_n4 = self.condb.pl_pur_rep(da_st_4, da_du_4)
        ws.cell(4, 5).value = self.pur_n4
        for n4 in range(len(self.item)):
            self.info4 = self.condb.pl_exp_rep(da_st_4, da_du_4, self.item[n4])
            ws.cell(e, 5).value = self.info4
            e = e + 1
        self.cn_n4 = self.condb.pl_cn_rep(da_st_4, da_du_4)
        ws.cell(14, 5).value = self.cn_n4
        self.bkin_n4 = self.condb.pl_exp_rep(da_st_4, da_du_4, self.item_bankint)
        ws.cell(16, 5).value = self.bkin_n4
        # Item_5
        f = 5
        da_st_5 = self.start_ed_5.text()
        da_du_5 = self.end_ed_5.text()
        ws.cell(2, 6).value = da_st_5 + "-" + da_du_5
        self.inv_n5 = self.condb.pl_inv_rep(da_st_5, da_du_5)
        ws.cell(3, 6).value = self.inv_n5
        self.pur_n5 = self.condb.pl_pur_rep(da_st_5, da_du_5)
        ws.cell(4, 6).value = self.pur_n5
        for n5 in range(len(self.item)):
            self.info5 = self.condb.pl_exp_rep(da_st_5, da_du_5, self.item[n5])
            ws.cell(f, 6).value = self.info5
            f = f + 1
        self.cn_n5 = self.condb.pl_cn_rep(da_st_5, da_du_5)
        ws.cell(14, 6).value = self.cn_n5
        self.bkin_n5 = self.condb.pl_exp_rep(da_st_5, da_du_5, self.item_bankint)
        ws.cell(16, 6).value = self.bkin_n5
        # Item_6
        g = 5
        da_st_6 = self.start_ed_6.text()
        da_du_6 = self.end_ed_6.text()
        ws.cell(2, 7).value = da_st_6 + "-" + da_du_6
        self.inv_n6 = self.condb.pl_inv_rep(da_st_6, da_du_6)
        ws.cell(3, 7).value = self.inv_n6
        self.pur_n6 = self.condb.pl_pur_rep(da_st_6, da_du_6)
        ws.cell(4, 7).value = self.pur_n6
        for n6 in range(len(self.item)):
            self.info6 = self.condb.pl_exp_rep(da_st_6, da_du_6, self.item[n6])
            ws.cell(g, 7).value = self.info6
            g = g + 1
        self.cn_n6 = self.condb.pl_cn_rep(da_st_6, da_du_6)
        ws.cell(14, 7).value = self.cn_n6
        self.bkin_n6 = self.condb.pl_exp_rep(da_st_6, da_du_6, self.item_bankint)
        ws.cell(16, 7).value = self.bkin_n6
        # Item_7
        h = 5
        da_st_7 = self.start_ed_7.text()
        da_du_7 = self.end_ed_7.text()
        ws.cell(2, 8).value = da_st_7 + "-" + da_du_7
        self.inv_n7 = self.condb.pl_inv_rep(da_st_7, da_du_7)
        ws.cell(3, 8).value = self.inv_n7
        self.pur_n7 = self.condb.pl_pur_rep(da_st_7, da_du_7)
        ws.cell(4, 8).value = self.pur_n7
        for n7 in range(len(self.item)):
            self.info7 = self.condb.pl_exp_rep(da_st_7, da_du_7, self.item[n7])
            ws.cell(h, 8).value = self.info7
            h = h + 1
        self.cn_n7 = self.condb.pl_cn_rep(da_st_7, da_du_7)
        ws.cell(14, 8).value = self.cn_n7
        self.bkin_n7 = self.condb.pl_exp_rep(da_st_7, da_du_7, self.item_bankint)
        ws.cell(16, 8).value = self.bkin_n7
        # Item_8
        i = 5
        da_st_8 = self.start_ed_8.text()
        da_du_8 = self.end_ed_8.text()
        ws.cell(2, 9).value = da_st_8 + "-" + da_du_8
        self.inv_n8 = self.condb.pl_inv_rep(da_st_8, da_du_8)
        ws.cell(3, 9).value = self.inv_n8
        self.pur_n8 = self.condb.pl_pur_rep(da_st_8, da_du_8)
        ws.cell(4, 9).value = self.pur_n8
        for n8 in range(len(self.item)):
            self.info8 = self.condb.pl_exp_rep(da_st_8, da_du_8, self.item[n8])
            ws.cell(i, 9).value = self.info8
            i = i + 1
        self.cn_n8 = self.condb.pl_cn_rep(da_st_8, da_du_8)
        ws.cell(14, 9).value = self.cn_n8
        self.bkin_n8 = self.condb.pl_exp_rep(da_st_8, da_du_8, self.item_bankint)
        ws.cell(16, 9).value = self.bkin_n8
        # Item_9
        j = 5
        da_st_9 = self.start_ed_9.text()
        da_du_9 = self.end_ed_9.text()
        ws.cell(2, 10).value = da_st_9 + "-" + da_du_9
        self.inv_n9 = self.condb.pl_inv_rep(da_st_9, da_du_9)
        ws.cell(3, 10).value = self.inv_n9
        self.pur_n9 = self.condb.pl_pur_rep(da_st_9, da_du_9)
        ws.cell(4, 10).value = self.pur_n9
        for n9 in range(len(self.item)):
            self.info9 = self.condb.pl_exp_rep(da_st_9, da_du_9, self.item[n9])
            ws.cell(j, 10).value = self.info9
            j = j + 1
        self.cn_n9 = self.condb.pl_cn_rep(da_st_9, da_du_9)
        ws.cell(14, 10).value = self.cn_n9
        self.bkin_n9 = self.condb.pl_exp_rep(da_st_9, da_du_9, self.item_bankint)
        ws.cell(16, 10).value = self.bkin_n9
        # Item_10
        k = 5
        da_st_10 = self.start_ed_10.text()
        da_du_10 = self.end_ed_10.text()
        ws.cell(2, 11).value = da_st_10 + "-" + da_du_10
        self.inv_n10 = self.condb.pl_inv_rep(da_st_10, da_du_10)
        ws.cell(3, 11).value = self.inv_n10
        self.pur_n10 = self.condb.pl_pur_rep(da_st_10, da_du_10)
        ws.cell(4, 11).value = self.pur_n10
        for n10 in range(len(self.item)):
            self.info10 = self.condb.pl_exp_rep(da_st_10, da_du_10, self.item[n10])
            ws.cell(k, 11).value = self.info10
            k = k + 1
        self.cn_n10 = self.condb.pl_cn_rep(da_st_10, da_du_10)
        ws.cell(14, 11).value = self.cn_n10
        self.bkin_n10 = self.condb.pl_exp_rep(da_st_10, da_du_10, self.item_bankint)
        ws.cell(16, 11).value = self.bkin_n10
        # Item_11
        la = 5
        da_st_11 = self.start_ed_11.text()
        da_du_11 = self.end_ed_11.text()
        ws.cell(2, 12).value = da_st_11 + "-" + da_du_11
        self.inv_n11 = self.condb.pl_inv_rep(da_st_11, da_du_11)
        ws.cell(3, 12).value = self.inv_n11
        self.pur_n11 = self.condb.pl_pur_rep(da_st_11, da_du_11)
        ws.cell(4, 12).value = self.pur_n11
        for n11 in range(len(self.item)):
            self.info11 = self.condb.pl_exp_rep(da_st_11, da_du_11, self.item[n11])
            ws.cell(la, 12).value = self.info11
            la = la + 1
        self.cn_n11 = self.condb.pl_cn_rep(da_st_11, da_du_11)
        ws.cell(14, 12).value = self.cn_n11
        self.bkin_n11 = self.condb.pl_exp_rep(da_st_11, da_du_11, self.item_bankint)
        ws.cell(16, 12).value = self.bkin_n11
        # Item_12
        m = 5
        da_st_12 = self.start_ed_12.text()
        da_du_12 = self.end_ed_12.text()
        ws.cell(2, 13).value = da_st_12 + "-" + da_du_12
        self.inv_n12 = self.condb.pl_inv_rep(da_st_12, da_du_12)
        ws.cell(3, 13).value = self.inv_n12
        self.pur_n12 = self.condb.pl_pur_rep(da_st_12, da_du_12)
        ws.cell(4, 13).value = self.pur_n12
        for n12 in range(len(self.item)):
            self.info12 = self.condb.pl_exp_rep(da_st_12, da_du_12, self.item[n12])
            ws.cell(m, 13).value = self.info12
            m = m + 1
        self.cn_n12 = self.condb.pl_cn_rep(da_st_12, da_du_12)
        ws.cell(14, 13).value = self.cn_n12
        self.bkin_n12 = self.condb.pl_exp_rep(da_st_12, da_du_12, self.item_bankint)
        ws.cell(16, 13).value = self.bkin_n12

        self.dirpath, _ = QFileDialog.getSaveFileName(self, "Save file",
                                                      "./" + "Profìt and Loss Statement" + self.start_ed.text() + "-" + self.end_ed_12.text(),
                                                      'xlsx(*.xlsx)')
        if self.dirpath != "":
            try:
                with open(self.dirpath, 'w'):
                    print(self.dirpath)
                    wb.save(self.dirpath)
                    QMessageBox.about(self, "About", "Done. !", )
            except IOError:
                QMessageBox.about(self, "About", "Pleare Close Document or Change Document Name. !", )
            else:
                pass
        else:
            QMessageBox.about(self, "About", "EXIT. !", )

    def reportpl_bu(self):
        self.Search_bu.clicked.connect(self.search_dl)

    def closeEvent(self, event):
        self.condb.closedb()
        print('window close')


if __name__ == "__main__":
    app = QApplication(sys.argv)
    win = Callplreport()
    win.show()
    sys.exit(app.exec_())
