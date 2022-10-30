# -*- coding: utf-8 -*-

import sys

from PyQt5 import QtCore
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QMessageBox
from return_listui import Ui_ret_list
from connectdb import database
import decimal
from PyQt5.QtCore import Qt
from openpyxl import load_workbook


class callret_list(QMainWindow, Ui_ret_list):
    my_Signal = QtCore.pyqtSignal()

    def __init__(self, list_info, parent=None):
        super(callret_list, self).__init__(parent)
        self.data_info = list_info
        self.data_inv = self.data_info[1]
        decimal.getcontext().rounding = "ROUND_HALF_UP"
        self.setupUi(self)
        self.condb = database()
        self.info = self.condb.display_cominfo()
        self.setWindowTitle('Retuen NOTE')
        self.setWindowModality(Qt.ApplicationModal)
        self.tolist()
        self.bu_click()

    def tolist(self):
        self.in_li.setText(self.data_info[1])
        self.da_li.setText(self.data_info[2])
        self.billt_li.setText(self.data_info[3])
        self.addr_li.setText(self.data_info[4])
        self.attn_li.setText(self.data_info[5])
        self.tel_lu.setText(self.data_info[6])
        self.cusid_li.setText(self.data_info[7])
        self.dued_li.setText(self.data_info[8])
        self.sales_li.setText(self.data_info[9])
        self.saleste_li.setText(self.data_info[10])
        self.p1.setText(self.data_info[11])
        self.p1m.setText(self.data_info[12])
        self.p1c.setText(self.data_info[13])
        self.p1p.setText(self.data_info[14])
        self.p2.setText(self.data_info[15])
        self.p2m.setText(self.data_info[16])
        self.p2c.setText(self.data_info[17])
        self.p2p.setText(self.data_info[18])
        self.p3.setText(self.data_info[19])
        self.p3m.setText(self.data_info[20])
        self.p3c.setText(self.data_info[21])
        self.p3p.setText(self.data_info[22])
        self.p4.setText(self.data_info[23])
        self.p4m.setText(self.data_info[24])
        self.p4c.setText(self.data_info[25])
        self.p4p.setText(self.data_info[26])
        self.p5.setText(self.data_info[27])
        self.p5m.setText(self.data_info[28])
        self.p5c.setText(self.data_info[29])
        self.p5p.setText(self.data_info[30])
        self.p6.setText(self.data_info[31])
        self.p6m.setText(self.data_info[32])
        self.p6c.setText(self.data_info[33])
        self.p6p.setText(self.data_info[34])
        self.p7.setText(self.data_info[35])
        self.p7m.setText(self.data_info[36])
        self.p7c.setText(self.data_info[37])
        self.p7p.setText(self.data_info[38])
        self.p8.setText(self.data_info[39])
        self.p8m.setText(self.data_info[40])
        self.p8c.setText(self.data_info[41])
        self.p8p.setText(self.data_info[42])
        self.p9.setText(self.data_info[43])
        self.p9m.setText(self.data_info[44])
        self.p9c.setText(self.data_info[45])
        self.p9p.setText(self.data_info[46])
        self.p10.setText(self.data_info[47])
        self.p10m.setText(self.data_info[48])
        self.p10c.setText(self.data_info[49])
        self.p10p.setText(self.data_info[50])
        self.p11.setText(self.data_info[51])
        self.p11m.setText(self.data_info[52])
        self.p11c.setText(self.data_info[53])
        self.p11p.setText(self.data_info[54])
        self.p12.setText(self.data_info[55])
        self.p12m.setText(self.data_info[56])
        self.p12c.setText(self.data_info[57])
        self.p12p.setText(self.data_info[58])
        self.p13.setText(self.data_info[59])
        self.p13m.setText(self.data_info[60])
        self.p13c.setText(self.data_info[61])
        self.p13p.setText(self.data_info[62])
        self.p14.setText(self.data_info[63])
        self.p14m.setText(self.data_info[64])
        self.p14c.setText(self.data_info[65])
        self.p14p.setText(self.data_info[66])
        self.p15.setText(self.data_info[67])
        self.p15m.setText(self.data_info[68])
        self.p15c.setText(self.data_info[69])
        self.p15p.setText(self.data_info[70])
        self.p16.setText(self.data_info[71])
        self.p16m.setText(self.data_info[72])
        self.p16c.setText(self.data_info[73])
        self.p16p.setText(self.data_info[74])
        self.p17.setText(self.data_info[75])
        self.p17m.setText(self.data_info[76])
        self.p17c.setText(self.data_info[77])
        self.p17p.setText(self.data_info[78])
        self.p18.setText(self.data_info[79])
        self.p18m.setText(self.data_info[80])
        self.p18c.setText(self.data_info[81])
        self.p18p.setText(self.data_info[82])
        self.p19.setText(self.data_info[83])
        self.p19m.setText(self.data_info[84])
        self.p19c.setText(self.data_info[85])
        self.p19p.setText(self.data_info[86])
        self.p20.setText(self.data_info[87])
        self.p20m.setText(self.data_info[88])
        self.p20c.setText(self.data_info[89])
        self.p20p.setText(self.data_info[90])
        self.c1.setText(self.data_info[91])
        self.c2.setText(self.data_info[92])
        self.c3.setText(self.data_info[93])
        self.to_am_ed.setText(self.data_info[94])

    def del_ret_bu(self):
        reply = QMessageBox.information(self, "Delete INVOICE", "Once deleted , Invoice cannot be recovered!",
                                        QMessageBox.Yes | QMessageBox.Close, QMessageBox.Close)
        if reply == QMessageBox.Yes:
            self.iv = self.data_inv
            self.invdata = self.condb.select_ret(self.iv)
            self.firdata = self.invdata[11:91]
            self.reindata = [i for i in self.firdata if i != '']
            print(self.reindata)
            self.exc_res = self.condb.excision_results(self.reindata,4)
            self.condb.delret_stock(self.exc_res)
            self.condb.ret_db(self.iv)
            self.close()


    def xls(self):
        wb = load_workbook('./tmp/Sample.xlsx')
        ws = wb['Invoice']
        ws['F1'] = "Return Note"
        ws['A1'] = self.info[1]
        ws['B2'] = self.info[2]
        ws['B3'] = self.sales_li.text()
        ws['B4'] = self.saleste_li.text()
        ws['B8'] = self.billt_li.text()
        ws['B9'] = self.addr_li.text()
        ws['B10'] = self.attn_li.text()
        ws['B11'] = self.tel_lu.text()
        ws['G3'] = self.da_li.text()
        ws['G4'] = self.in_li.text()
        ws['G5'] = self.cusid_li.text()
        ws['G6'] = self.dued_li.text()
        ws['A16'] = self.p1.text()
        ws['A17'] = self.p2.text()
        ws['A18'] = self.p3.text()
        ws['A19'] = self.p4.text()
        ws['A20'] = self.p5.text()
        ws['A21'] = self.p6.text()
        ws['A22'] = self.p7.text()
        ws['A23'] = self.p8.text()
        ws['A24'] = self.p9.text()
        ws['A25'] = self.p10.text()
        ws['A26'] = self.p11.text()
        ws['A27'] = self.p12.text()
        ws['A28'] = self.p13.text()
        ws['A29'] = self.p14.text()
        ws['A30'] = self.p15.text()
        ws['A31'] = self.p16.text()
        ws['A32'] = self.p17.text()
        ws['A33'] = self.p18.text()
        ws['A34'] = self.p19.text()
        ws['A35'] = self.p20.text()
        ws['B16'] = self.p1m.text()
        ws['B17'] = self.p2m.text()
        ws['B18'] = self.p3m.text()
        ws['B19'] = self.p4m.text()
        ws['B20'] = self.p5m.text()
        ws['B21'] = self.p6m.text()
        ws['B22'] = self.p7m.text()
        ws['B23'] = self.p8m.text()
        ws['B24'] = self.p9m.text()
        ws['B25'] = self.p10m.text()
        ws['B26'] = self.p11m.text()
        ws['B27'] = self.p12m.text()
        ws['B28'] = self.p13m.text()
        ws['B29'] = self.p14m.text()
        ws['B30'] = self.p15m.text()
        ws['B31'] = self.p16m.text()
        ws['B32'] = self.p17m.text()
        ws['B33'] = self.p18m.text()
        ws['B34'] = self.p19m.text()
        ws['B35'] = self.p20m.text()
        ws['A41'] = self.c1.text()
        ws['A42'] = self.c2.text()
        ws['A43'] = self.c3.text()

        ###Product
        self.toam = []
        self.toqty = []
        if self.p1c.text() != "":
            ws['E16'] = decimal.Decimal(self.p1c.text())
            ws['F16'] = decimal.Decimal(self.p1p.text())
            ws['G16'] = decimal.Decimal(self.p1p.text()) * decimal.Decimal(self.p1c.text())
            self.toam.append(ws['G16'].value)
            self.toqty.append(ws['E16'].value)
        if self.p2c.text() != "":
            ws['E17'] = decimal.Decimal(self.p2c.text())
            ws['F17'] = decimal.Decimal(self.p2p.text())
            ws['G17'] = decimal.Decimal(self.p2p.text()) * decimal.Decimal(self.p2c.text())
            self.toam.append(ws['G17'].value)
            self.toqty.append(ws['E17'].value)
        if self.p3c.text() != "":
            ws['E18'] = decimal.Decimal(self.p3c.text())
            ws['F18'] = decimal.Decimal(self.p3p.text())
            ws['G18'] = decimal.Decimal(self.p3p.text()) * decimal.Decimal(self.p3c.text())
            self.toam.append(ws['G18'].value)
            self.toqty.append(ws['E18'].value)
        if self.p4c.text() != "":
            ws['E19'] = decimal.Decimal(self.p4c.text())
            ws['F19'] = decimal.Decimal(self.p4p.text())
            ws['G19'] = decimal.Decimal(self.p4p.text()) * decimal.Decimal(self.p4c.text())
            self.toam.append(ws['G19'].value)
            self.toqty.append(ws['E19'].value)
        if self.p5c.text() != "":
            ws['E20'] = decimal.Decimal(self.p5c.text())
            ws['F20'] = decimal.Decimal(self.p5p.text())
            ws['G20'] = decimal.Decimal(self.p5p.text()) * decimal.Decimal(self.p5c.text())
            self.toam.append(ws['G20'].value)
            self.toqty.append(ws['E20'].value)
        if self.p6c.text() != "":
            ws['E21'] = decimal.Decimal(self.p6c.text())
            ws['F21'] = decimal.Decimal(self.p6p.text())
            ws['G21'] = decimal.Decimal(self.p6p.text()) * decimal.Decimal(self.p6c.text())
            self.toam.append(ws['G21'].value)
            self.toqty.append(ws['E21'].value)
        if self.p7c.text() != "":
            ws['E22'] = decimal.Decimal(self.p7c.text())
            ws['F22'] = decimal.Decimal(self.p7p.text())
            ws['G22'] = decimal.Decimal(self.p7p.text()) * decimal.Decimal(self.p7c.text())
            self.toam.append(ws['G22'].value)
            self.toqty.append(ws['E22'].value)
        if self.p8c.text() != "":
            ws['E23'] = decimal.Decimal(self.p8c.text())
            ws['F23'] = decimal.Decimal(self.p8p.text())
            ws['G23'] = decimal.Decimal(self.p8p.text()) * decimal.Decimal(self.p8c.text())
            self.toam.append(ws['G23'].value)
            self.toqty.append(ws['E23'].value)
        if self.p9c.text() != "":
            ws['E24'] = decimal.Decimal(self.p9c.text())
            ws['F24'] = decimal.Decimal(self.p9p.text())
            ws['G24'] = decimal.Decimal(self.p9p.text()) * decimal.Decimal(self.p9c.text())
            self.toam.append(ws['G24'].value)
            self.toqty.append(ws['E24'].value)
        if self.p10c.text() != "":
            ws['E25'] = decimal.Decimal(self.p10c.text())
            ws['F25'] = decimal.Decimal(self.p10p.text())
            ws['G25'] = decimal.Decimal(self.p10p.text()) * decimal.Decimal(self.p10c.text())
            self.toam.append(ws['G25'].value)
            self.toqty.append(ws['E25'].value)
        if self.p11c.text() != "":
            ws['E26'] = decimal.Decimal(self.p11c.text())
            ws['F26'] = decimal.Decimal(self.p11p.text())
            ws['G26'] = decimal.Decimal(self.p11p.text()) * decimal.Decimal(self.p11c.text())
            self.toam.append(ws['G26'].value)
            self.toqty.append(ws['E26'].value)
        if self.p12c.text() != "":
            ws['E27'] = decimal.Decimal(self.p12c.text())
            ws['F27'] = decimal.Decimal(self.p12p.text())
            ws['G27'] = decimal.Decimal(self.p12p.text()) * decimal.Decimal(self.p12c.text())
            self.toam.append(ws['G27'].value)
            self.toqty.append(ws['E27'].value)
        if self.p13c.text() != "":
            ws['E28'] = decimal.Decimal(self.p13c.text())
            ws['F28'] = decimal.Decimal(self.p13p.text())
            ws['G28'] = decimal.Decimal(self.p13p.text()) * decimal.Decimal(self.p13c.text())
            self.toam.append(ws['G28'].value)
            self.toqty.append(ws['E28'].value)
        if self.p14c.text() != "":
            ws['E29'] = decimal.Decimal(self.p14c.text())
            ws['F29'] = decimal.Decimal(self.p14p.text())
            ws['G29'] = decimal.Decimal(self.p14p.text()) * decimal.Decimal(self.p14c.text())
            self.toam.append(ws['G29'].value)
            self.toqty.append(ws['E29'].value)
        if self.p15c.text() != "":
            ws['E30'] = decimal.Decimal(self.p15c.text())
            ws['F30'] = decimal.Decimal(self.p15p.text())
            ws['G30'] = decimal.Decimal(self.p15p.text()) * decimal.Decimal(self.p15c.text())
            self.toam.append(ws['G30'].value)
            self.toqty.append(ws['E30'].value)
        if self.p16c.text() != "":
            ws['E31'] = decimal.Decimal(self.p16c.text())
            ws['F31'] = decimal.Decimal(self.p16p.text())
            ws['G31'] = decimal.Decimal(self.p16p.text()) * decimal.Decimal(self.p16c.text())
            self.toam.append(ws['G31'].value)
            self.toqty.append(ws['E31'].value)
        if self.p17c.text() != "":
            ws['E32'] = decimal.Decimal(self.p17c.text())
            ws['F32'] = decimal.Decimal(self.p17p.text())
            ws['G32'] = decimal.Decimal(self.p17p.text()) * decimal.Decimal(self.p17c.text())
            self.toam.append(ws['G32'].value)
            self.toqty.append(ws['E32'].value)
        if self.p18c.text() != "":
            ws['E33'] = decimal.Decimal(self.p18c.text())
            ws['F33'] = decimal.Decimal(self.p18p.text())
            ws['G33'] = decimal.Decimal(self.p18p.text()) * decimal.Decimal(self.p18c.text())
            self.toam.append(ws['G33'].value)
            self.toqty.append(ws['E33'].value)
        if self.p19c.text() != "":
            ws['E34'] = decimal.Decimal(self.p19c.text())
            ws['F34'] = decimal.Decimal(self.p19p.text())
            ws['G34'] = decimal.Decimal(self.p19p.text()) * decimal.Decimal(self.p19c.text())
            self.toam.append(ws['G34'].value)
            self.toqty.append(ws['E34'].value)
        if self.p20c.text() != "":
            ws['E35'] = decimal.Decimal(self.p20c.text())
            ws['F35'] = decimal.Decimal(self.p20p.text())
            ws['G35'] = decimal.Decimal(self.p20p.text()) * decimal.Decimal(self.p20c.text())
            self.toam.append(ws['G35'].value)
            self.toqty.append(ws['E35'].value)
        print(self.toam)
        ws['G38'] = sum(self.toam)
        print(self.toqty)
        ws['E38'] = sum(self.toqty)
        self.dirpath, _ = QFileDialog.getSaveFileName(self, "Save file", "./" + self.in_li.text(), 'xlsx(*.xlsx)')
        if self.dirpath != "":
            try:
                with open(self.dirpath, 'w'):
                    print(self.dirpath)
                    wb.save(self.dirpath)
                    QMessageBox.about(self, "About", "Done. !", )
            except IOError:
                QMessageBox.about(self, "About", "Pleare Close Document or Change Document Name. !", )
            else:
                pass
        else:
            QMessageBox.about(self, "About", "EXIT. !", )


    def checkBoxChangedAction(self, state):
        if state == Qt.Checked:
            self.da_li.setEnabled(True)
            self.dued_li.setEnabled(True)
        else:
            self.da_li.setEnabled(False)
            self.dued_li.setEnabled(False)

    def da_up_bu(self):
        reply = QMessageBox.information(self, "UPLOAD Date", "Once UPLOAD , Date cannot be recovered!",
                                        QMessageBox.Yes | QMessageBox.Close, QMessageBox.Close)
        if reply == QMessageBox.Yes:
            da = self.da_li.text()
            dua = self.dued_li.text()
            inv = self.in_li.text()
            self.condb.up_ret_data(da,dua,inv)
            self.close()

    def bu_click(self):
        self.del_bu.clicked.connect(self.del_ret_bu)
        self.print_bu.clicked.connect(self.xls)
        self.data_checkBox.stateChanged.connect(self.checkBoxChangedAction)
        self.data_upload.clicked.connect(self.da_up_bu)

    def sendEditContent(self):
        self.my_Signal.emit()

    def closeEvent(self, event):
        self.sendEditContent()
        self.condb.closedb()
        print('window close')


if __name__ == "__main__":
    app = QApplication(sys.argv)
    win = callret_list()
    win.show()
    sys.exit(app.exec_())
